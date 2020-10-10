from openpyxl import Workbook, load_workbook

wb = load_workbook('dataset.xlsx')
#wb.create_sheet('Motivos Isolamento')
sheet = wb['Sheet2']
print(sheet.cell(row = 2, column = 1).value)
input()
